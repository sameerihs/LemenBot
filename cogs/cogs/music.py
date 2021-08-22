import discord, datetime, time
from discord.ext import commands
import asyncio
import functools
import itertools
import math
import random
import asyncpg
import ast
from collections import deque
import youtube_dl
from async_timeout import timeout
import discord
import os
import discord.utils
from discord.ext.commands import has_permissions, MissingPermissions
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
from gtts import gTTS
from tempfile import TemporaryFile
from discord.utils import get
from discord_components import DiscordComponents, Button, ButtonStyle, InteractionType, component
import pygsheets
from dpymenus import Page, TextMenu
from dpymenus.constants import CONFIRM
from dislash import SlashClient, SelectMenu, SelectOption
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side

file = load_workbook(r"C:\SAMEER MAIN\Sameer\Code\Openpyxl\data.xlsx")
sheet = file["MAIN"]
fonttemplate = Font(name='Calibre',size=12,color='FF000000')



youtube_dl.utils.bug_reports_message = lambda: ''


class VoiceError(Exception):
    pass


class YTDLError(Exception):
    pass


class YTDLSource(discord.PCMVolumeTransformer):
    YTDL_OPTIONS = {
        'format': 'bestaudio/best',
        'extractaudio': True,
        'audioformat': 'mp3',
        'outtmpl': '%(extractor)s-%(id)s-%(title)s.%(ext)s',
        'restrictfilenames': True,
        'noplaylist': True,
        'nocheckcertificate': True,
        'ignoreerrors': False,
        'logtostderr': False,
        'quiet': True,
        'no_warnings': True,
        'default_search': 'auto',
        'source_address': '0.0.0.0',
    }

    FFMPEG_OPTIONS = {
        'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5',
        'options': '-vn',
    }

    ytdl = youtube_dl.YoutubeDL(YTDL_OPTIONS)

    def __init__(self, ctx: commands.Context, source: discord.FFmpegPCMAudio, *, data: dict, volume: float = 50):
        super().__init__(source, volume)

        self.requester = ctx.author
        self.channel = ctx.channel
        self.data = data

        self.uploader = data.get('uploader')
        self.uploader_url = data.get('uploader_url')
        date = data.get('upload_date')
        self.upload_date = date[6:8] + '.' + date[4:6] + '.' + date[0:4]
        self.title = data.get('title')
        self.thumbnail = data.get('thumbnail')
        self.description = data.get('description')
        self.duration = self.parse_duration(int(data.get('duration')))
        self.tags = data.get('tags')
        self.url = data.get('webpage_url')
        self.views = data.get('view_count')
        self.likes = data.get('like_count')
        self.dislikes = data.get('dislike_count')
        self.stream_url = data.get('url')

    def __str__(self):
        return '**{0.title}** by **{0.uploader}**'.format(self)

    @classmethod
    async def create_source(cls, ctx: commands.Context, search: str, *, loop: asyncio.BaseEventLoop = None):
        loop = loop or asyncio.get_event_loop()

        partial = functools.partial(cls.ytdl.extract_info, search, download=False, process=False)
        data = await loop.run_in_executor(None, partial)

        if data is None:
            raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        if 'entries' not in data:
            process_info = data
        else:
            process_info = None
            for entry in data['entries']:
                if entry:
                    process_info = entry
                    break

            if process_info is None:
                raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        webpage_url = process_info['webpage_url']
        partial = functools.partial(cls.ytdl.extract_info, webpage_url, download=False)
        processed_info = await loop.run_in_executor(None, partial)

        if processed_info is None:
            raise YTDLError('Couldn\'t fetch `{}`'.format(webpage_url))

        if 'entries' not in processed_info:
            info = processed_info
        else:
            info = None
            while info is None:
                try:
                    info = processed_info['entries'].pop(0)
                except IndexError:
                    raise YTDLError('Couldn\'t retrieve any matches for `{}`'.format(webpage_url))

        return cls(ctx, discord.FFmpegPCMAudio(info['url'], **cls.FFMPEG_OPTIONS,executable=r"C:\SAMEER MAIN\Sameer\Code\Lemen Bot\LemenBot\ffmpeg\bin\ffmpeg.exe"), data=info)

    @staticmethod
    def parse_duration(duration: int):
        minutes, seconds = divmod(duration, 60)
        hours, minutes = divmod(minutes, 60)
        days, hours = divmod(hours, 24)

        duration = []
        if days > 0:
            duration.append('{} days'.format(days))
        if hours > 0:
            duration.append('{} hours'.format(hours))
        if minutes > 0:
            duration.append('{} minutes'.format(minutes))
        if seconds > 0:
            duration.append('{} seconds'.format(seconds))

        return ', '.join(duration)



class Song:
    __slots__ = ('source', 'requester')

    def __init__(self, source: YTDLSource):
        self.source = source
        self.requester = source.requester

    def create_embed(self):
        embed = (discord.Embed(title='Now playing',
                               description='```py\n{0.source.title}\n```'.format(self),
                               color=0xfff44f)
                 .add_field(name='Duration', value=self.source.duration)
                 .add_field(name='Requested by', value=self.requester.mention)
                 .add_field(name='Uploader', value='[{0.source.uploader}]({0.source.uploader_url})'.format(self))
                 .add_field(name='URL', value='[Click]({0.source.url})'.format(self))
                 .set_thumbnail(url=self.source.thumbnail))

        return embed


class SongQueue(asyncio.Queue):
    def __getitem__(self, item):
        if isinstance(item, slice):
            return list(itertools.islice(self._queue, item.start, item.stop, item.step))
        else:
            return self._queue[item]

    def __iter__(self):
        return self._queue.__iter__()

    def __len__(self):
        return self.qsize()

    def clear(self):
        self._queue.clear()
    
    def shuffle(self):
        random.shuffle(self._queue)

    def remove(self, index: int):
        del self._queue[index]


class VoiceState:
    def __init__(self, bot: commands.Bot, ctx: commands.Context):
        self.bot = bot
        self._ctx = ctx

        self.current = None
        self.voice = None
        self.next = asyncio.Event()
        self.songs = SongQueue()

        self._loop = False
        self._volume = 0.5
        self.skip_votes = set()

        self.audio_player = bot.loop.create_task(self.audio_player_task())

    def __del__(self):
        self.audio_player.cancel()

    @property
    def loop(self):
        return self._loop

    @loop.setter
    def loop(self, value: bool):
        self._loop = value

    @property
    def volume(self):
        return self._volume

    @volume.setter
    def volume(self, value: float):
        self._volume = value

    @property
    def is_playing(self):
        return self.voice and self.current

    async def audio_player_task(self):
        while True:
            self.next.clear()

            if not self.loop:
                try:
                    async with timeout(360):
                        self.current = await self.songs.get()

                except asyncio.TimeoutError:
                    self.bot.loop.create_task(self.stop())
                    

                    return

            self.current.source.volume = self._volume
            self.voice.play(self.current.source, after=self.play_next_song)
            await self.current.source.channel.send(embed=self.current.create_embed())

            await self.next.wait()

    def play_next_song(self, error=None):
        if error:
            raise VoiceError(str(error))

        self.next.set()

    def skip(self):
        self.skip_votes.clear()

        if self.is_playing:
            self.voice.stop()

    async def stop(self):
        self.songs.clear()

        if self.voice:
            await self.voice.disconnect()
            self.voice = None


class Music(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.voice_states = {}

    def get_voice_state(self, ctx: commands.Context):
        state = self.voice_states.get(ctx.guild.id)
        if not state:
            state = VoiceState(self.bot, ctx)
            self.voice_states[ctx.guild.id] = state

        return state

    def cog_unload(self):
        for state in self.voice_states.values():
            self.bot.loop.create_task(state.stop())

    def cog_check(self, ctx: commands.Context):
        if not ctx.guild:
            raise commands.NoPrivateMessage('This command can\'t be used in DM channels.')

        return True

    async def cog_before_invoke(self, ctx: commands.Context):
        ctx.voice_state = self.get_voice_state(ctx)

    async def cog_command_error(self, ctx: commands.Context, error: commands.CommandError):
        print ('An error occurred: {}'.format(str(error)))

    @commands.command()
    async def _join(self, ctx: commands.Context):
        """Joins a voice channel."""

        destination = ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(name='summon',aliases=['join'])
    #@commands.has_permissions(manage_guild=True)
    async def _summon(self, ctx: commands.Context, *, channel: discord.VoiceChannel = None):
        """Summons the bot to a voice channel.
        If no channel was specified, it joins your channel.
        """

        if not channel and not ctx.author.voice:
            await ctx.send('```You are neither connected to a voice channel nor specified a channel to join.```')

        destination = channel or ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(aliases=['leave','disconnect','exit'])
    async def _leave(self, ctx: commands.Context):
        await asyncio.sleep(0.5)
        await ctx.voice_state.stop()
        del self.voice_states[ctx.guild.id]
        if not ctx.voice_state.voice:
            return await ctx.send('Not connected to any voice channel.')
        await ctx.voice_state.stop()
        del self.voice_states[ctx.guild.id]
        #try:

        #await ctx.voice_state.stop()
        #del self.voice_states[ctx.guild.id]
        ##await ctx.voice_client.disconnect()
        #await ctx.send("Leaving...")
        #return

        #except(TypeError, AttributeError):
        #   await ctx.send("Can't disconnect from a voice channel when I'm not in one")
        #   return
#
        #if not ctx.voice_state.voice:
        #    return await ctx.send('Not connected to any voice channel.')
#
        #await ctx.voice_state.stop()
        #del self.voice_states[ctx.guild.id]




    @commands.command(aliases=['vl','v'])
    async def volume(self, ctx, volume: int):
        """Changes the player's volume"""

        if ctx.voice_client is None:
            return await ctx.send("Not connected to a voice channel.")

        ctx.voice_client.source.volume = volume / 100
        if volume<=0:
            await ctx.send("⚪⚪⚪⚪⚪⚪⚪⚪⚪⚪ {}%".format(volume))
        if 1<=volume<=10:
            await ctx.send("🟡⚪⚪⚪⚪⚪⚪⚪⚪⚪ {}%".format(volume))
        if  10<volume<=20:
            await ctx.send("🟡🟡⚪⚪⚪⚪⚪⚪⚪⚪ {}%".format(volume))
        if  20<volume<=30:
            await ctx.send("🟡🟡🟡⚪⚪⚪⚪⚪⚪⚪ {}%".format(volume))
        if  30<volume<=40:
            await ctx.send("🟡🟡🟡🟡⚪⚪⚪⚪⚪⚪ {}%".format(volume))
        if  40<volume<=50:
            await ctx.send("🟡🟡🟡🟡🟡⚪⚪⚪⚪⚪ {}%".format(volume))
        if  50<volume<=60:
            await ctx.send("🟡🟡🟡🟡🟡🟡⚪⚪⚪⚪ {}%".format(volume))
        if  60<volume<=70:
            await ctx.send("🟡🟡🟡🟡🟡🟡🟡⚪⚪⚪ {}%".format(volume))
        if  70<volume<=80:
            await ctx.send("🟡🟡🟡🟡🟡🟡🟡🟡⚪⚪ {}%".format(volume))
        if  80<volume<=90:
            await ctx.send("🟡🟡🟡🟡🟡🟡🟡🟡🟡⚪ {}%".format(volume))
        if  90<volume<=100:
            await ctx.send("🟡🟡🟡🟡🟡🟡🟡🟡🟡🟡 {}%".format(volume))
        if  volume>100:
            await ctx.send("🟡🟡🟡🟡🟡🟡🟡🟡🟡🟡 {}%".format(volume))








    @commands.command(name='now', aliases=['current', 'playing'])
    async def _now(self, ctx: commands.Context):
        """Displays the currently playing song."""

        await ctx.send(embed=ctx.voice_state.current.create_embed())

    @commands.command(name='pause')
    async def _pause(self, ctx: commands.Context):
        """Pauses the currently playing song."""

        if ctx.voice_state.voice.is_playing():
            ctx.voice_state.voice.pause()
            await ctx.message.add_reaction('⏯')

    @commands.command(name='resume')                        
    async def _resume(self, ctx: commands.Context):
        """Resumes a currently paused song."""

        if ctx.voice_state.voice.is_paused():
            ctx.voice_state.voice.resume()
            await ctx.message.add_reaction('⏯')

    #@commands.command(name='stop')
    #async def _stop(self, ctx: commands.Context):
    #    await asyncio.sleep(0.5)
    #    if not ctx.voice_state.voice:
    #        return await ctx.send('Not connected to any voice channel.')
    #    if  ctx.voice_state.voice:
    #        await ctx.voice_state.stop()
    #        del self.voice_states[ctx.guild.id]
    #    await ctx.voice_state.stop()
    #    del self.voice_states[ctx.guild.id]
        


    @commands.command(name='skip')
    async def _skip(self, ctx: commands.Context):
        if ctx.voice_state.is_playing:
            await ctx.message.add_reaction('⏭')
            ctx.voice_state.skip()
    



    @commands.command(name='queue',aliases=['q'])
    async def _queue(self, ctx: commands.Context, *, page: int = 1):
        """Shows the player's queue.
        You can optionally specify the page to show. Each page contains 10 elements.
        """

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Empty queue.')

        items_per_page = 10
        pages = math.ceil(len(ctx.voice_state.songs) / items_per_page)

        start = (page - 1) * items_per_page
        end = start + items_per_page

        queue = ''
        for i, song in enumerate(ctx.voice_state.songs[start:end], start=start):
            queue += '`{0}.` [**{1.source.title}**]({1.source.url})\n'.format(i + 1, song)

        embed = (discord.Embed(description='**{} tracks:**\n\n{}'.format(len(ctx.voice_state.songs), queue))
                 .set_footer(text='Viewing page {}/{}'.format(page, pages)))
        await ctx.send(embed=embed)

    @commands.command(name='shuffle')
    async def _shuffle(self, ctx: commands.Context):
        """Shuffles the queue."""

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Empty queue.')

        ctx.voice_state.songs.shuffle()
        await ctx.message.add_reaction('✅')

    @commands.command(name='remove')
    async def _remove(self, ctx: commands.Context, index: int):
        """Removes a song from the queue at a given index."""

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Empty queue.')

        ctx.voice_state.songs.remove(index - 1)
        await ctx.message.add_reaction('✅')

    @commands.command(name='loop')
    async def _loop(self, ctx: commands.Context):
        """Loops the currently playing song.
        Invoke this command again to unloop the song.
        """

        if ctx.voice_state.is_playing:
            
            ctx.voice_state.loop = not ctx.voice_state.loop
        # Inverse boolean value to loop and unloop.
        ctx.voice_state.loop = not ctx.voice_state.loop
        await ctx.message.add_reaction('✅')

    @commands.command(name='play',aliases=['pl','p','Play','PLAY','P'])
    async def _play(self, ctx: commands.Context, *, search: str):
        if not ctx.voice_state.voice:
            await ctx.invoke(self._join)
        async with ctx.typing():
            try:
                source = await YTDLSource.create_source(ctx, search, loop=self.bot.loop)
            except YTDLError as e:
                await ctx.send('An error occurred while processing this request: {}'.format(str(e)))
            else:
                song = Song(source)
                await ctx.voice_state.songs.put(song)
                await ctx.send('Enqueued {}'.format(str(source)))



    @commands.command(name='sp',aliases=['spotifyplay'])
    async def _sp(self, ctx: commands.Context, *, link):
        client_id = "5f32d8d16a3b4ca58df7831a2d96f65f"
        client_secret = "a9f10697de40424786b8dcf2d5374652"
        client_credentials_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
        sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)
        main=[]
        main_artist=[]
        pop = sp.playlist(link)
        nn=len(pop['tracks']['items'])
        for j in range(0,nn-1):
            main.append(pop['tracks']['items'][j]['track']['name'])
            main_artist.append(pop['tracks']['items'][j]['track']['artists'][0]['name'])
        for ki in range(len(main)):
            rm=str(main[ki])+" by "+str(main_artist[ki])
            if not ctx.voice_state.voice:
                await ctx.invoke(self._join)
            try:
                source = await YTDLSource.create_source(ctx, rm, loop=self.bot.loop)
            except YTDLError as e:
                print('An error occurred while processing this request: {}'.format(str(e)))
            else:
                song = Song(source)
                await ctx.voice_state.songs.put(song)
                


    @commands.command()
    async def mysp(self, ctx):

        file = load_workbook(r"C:\SAMEER MAIN\Sameer\Code\Openpyxl\data.xlsx")
        sheet = file["MAIN"]
        fonttemplate = Font(name='Calibre',size=12,color='FF000000')

        a={}
        client_id = "5f32d8d16a3b4ca58df7831a2d96f65f"
        client_secret = "a9f10697de40424786b8dcf2d5374652"
        client_credentials_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
        sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)

        link_list=[]
        all_playlist_names=[]
        insideholder=""

        for i in range(1,100):
            d="D"+str(i)
            if sheet[d].value==str(ctx.author.display_name):
                f="F"+str(i)
                link_list.append(sheet[f].value)


        if len(link_list)==0:
            await ctx.send("You have no playlist saved. Use the **.add_playlist <link>** command to add a playlist")
        else:
            for l in link_list:
                pops = sp.playlist(l)
                playlist_name=pops['name']
                all_playlist_names.append(playlist_name)
    
            for _msa in all_playlist_names:
                rdmstr=str(random.randint(1,10000000))
                insideholder+="SelectOption(\""+_msa+"\",\""+rdmstr+"\""+"),"
    
    
            final_holder=eval("["+insideholder[0:-1]+"]")
            print(final_holder)
    
            holder=final_holder
            msg = await ctx.send("``Your saved playlists``",components=[SelectMenu(custom_id="test",placeholder="Select the playlist to play",max_values=1,options=holder)])
            while True:
                inter = await msg.wait_for_dropdown()
                labels = [option.label for option in inter.select_menu.selected_options]
                name12=all_playlist_names
                link12=link_list
                for ilm in range (len(name12)):
                    a[name12[ilm]] = link12[ilm]
                ff=str(labels[0])
                labels.pop(0)
                await inter.send("Adding {} to queue".format(ff),delete_after=5,ephemeral=True)
                main=[]
                main_artist=[]
                pop = sp.playlist(a[ff])
                nn=len(pop['tracks']['items'])
                for j in range(0,nn-1):
                    main.append(pop['tracks']['items'][j]['track']['name'])
                    main_artist.append(pop['tracks']['items'][j]['track']['artists'][0]['name'])
                for ki in range(len(main)):
                    rm=str(main[ki])+" by "+str(main_artist[ki])
                    if not ctx.voice_state.voice:
                        await ctx.invoke(self._join)
                    try:
                        source = await YTDLSource.create_source(ctx, rm, loop=self.bot.loop)
                    except YTDLError as e:
                        print('An error occurred while processing this request: {}'.format(str(e)))
                    else:
                        song = Song(source)
                        await ctx.voice_state.songs.put(song)
            
        link_list.clear()
        all_playlist_names.clear()
        insideholder=""
























    #@commands.command()
    #async def mysp(self, ctx):
    #    client1 = pygsheets.authorize(service_file='./client_secret.json')
    #    sheet = client1.open('Lemen Control Panel')
    #    worksheet = sheet[1]
    #    myname=str(ctx.author.display_name)
    #    header = worksheet.find("#")
    #    indexnum=int(str(header[0].label)[-1])
    #    nny=[]
    #    linkurl=[]
    #    playlistdict={}
    #    for i in range(indexnum):
    #        index="B"+str(i)
    #        checkname=str(worksheet.get_value(index))
    #        if myname==checkname:
    #            printer=str(worksheet.get_value("D"+str(i)))
    #            linkurl.append(printer)
    #            import spotipy
    #            from spotipy.oauth2 import SpotifyClientCredentials
    #            client_id = "5f32d8d16a3b4ca58df7831a2d96f65f"
    #            client_secret = "a9f10697de40424786b8dcf2d5374652"
    #            client_credentials_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
    #            sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)
    #            main=[]
    #            main_artist=[]
    #            pops = sp.playlist(printer)
    #            nny.append(pops['name'])
    #    for idekwhat in range(len(nny)):
    #        playlistdict[nny[idekwhat]] = linkurl[idekwhat]
    #    
#
    #      
    #    for j in nny:
    #        print(j)
    #        await ctx.send("Click on the Button to play the playlist",components = [Button(style= ButtonStyle.blue,label = j)])
#
#
    #    final_player=[]
    #    for k in nny:
    #        interaction = await self.bot.wait_for("button_click", check = lambda i: i.component.label.startswith(k))
    #        await interaction.respond(content = "Adding {} to queue...".format(k))
    #        final_player.append(playlistdict[k])
#
    #    for klh in final_player:
    #        client_id = "5f32d8d16a3b4ca58df7831a2d96f65f"
    #        client_secret = "a9f10697de40424786b8dcf2d5374652"
    #        client_credentials_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
    #        sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)
    #        main=[]
    #        main_artist=[]
    #        pop = sp.playlist(klh)
    #        nn=len(pop['tracks']['items'])
    #        for j in range(0,nn-1):
    #            main.append(pop['tracks']['items'][j]['track']['name'])
    #            main_artist.append(pop['tracks']['items'][j]['track']['artists'][0]['name'])
    #        for ki in range(len(main)):
    #            rm=str(main[ki])+" by "+str(main_artist[ki])
    #            if not ctx.voice_state.voice:
    #                await ctx.invoke(self._join)
    #            try:
    #                source = await YTDLSource.create_source(ctx, rm, loop=self.bot.loop)
    #            except YTDLError as e:
    #                print('An error occurred while processing this request: {}'.format(str(e)))
    #            else:
    #                song = Song(source)
    #                await ctx.voice_state.songs.put(song)
    #        #await asyncio.sleep(3)
            













    @commands.command()
    async def say(self, ctx: commands.Context, *, speech):
        message_queue = deque([])
        can_speak = True
        message = speech
        usernick = ctx.message.author.display_name
        message = message
        try:
            vc = ctx.message.guild.voice_client
            if not vc.is_playing():
                tts = gTTS(message)
                f = TemporaryFile()
                tts.write_to_fp(f)
                f.seek(0)
                vc.play(discord.FFmpegPCMAudio(f,pipe=True,executable=r"C:\SAMEER MAIN\Sameer\Code\Lemen Bot\LemenBot\ffmpeg\bin\ffmpeg.exe"))

            else:
                message_queue.append(message)
                while vc.is_playing():
                    await asyncio.sleep(0.1)
                tts = gTTS(message_queue.popleft())
                f = TemporaryFile()
                tts.write_to_fp(f)
                f.seek(0)
                vc.play(discord.FFmpegPCMAudio(f, pipe=True,executable=r"C:\SAMEER MAIN\Sameer\Code\Lemen Bot\LemenBot\ffmpeg\bin\ffmpeg.exe"))
        except(TypeError, AttributeError):
            try:
                tts = gTTS(message)
                f = TemporaryFile()
                tts.write_to_fp(f)
                f.seek(0)
                channel = ctx.message.author.voice.channel
                vc = await channel.connect()
                vc.play(discord.FFmpegPCMAudio(f, pipe=True,executable=r"C:\SAMEER MAIN\Sameer\Code\Lemen Bot\LemenBot\ffmpeg\bin\ffmpeg.exe"))
            except(AttributeError, TypeError):
                await ctx.send("I'm not in a voice channel and neither are you!")
            return
        f.close()


    @commands.command()
    @commands.has_permissions(manage_channels=True)
    async def vcannounce(self,ctx: commands.Context, *, name: str):
        main=[]
        await ctx.send("Select the VC's you'd like to announce:")
        for voice_channel in ctx.message.guild.voice_channels:
            vc_name="<#"+str(voice_channel.id)+">"
            msg_1=await ctx.send(vc_name)
            main.append(msg_1.id)
            await msg_1.add_reaction('✅')
        
        await asyncio.sleep(10)
        m_m="**The annoucement will be made in the following voice channels:**"
        await ctx.send(m_m)
        for a in main:
            message = await ctx.channel.fetch_message(a)
            reaction = get(message.reactions, emoji="✅")
            if reaction and reaction.count > 1:
                await ctx.send(message.content)
                destination =bot.get_channel(int(message.content[2: ].replace(">","")))
                if not ctx.voice_state.voice:
                    ctx.voice_state.voice = await destination.connect()
                    #await ctx.voice_state.voice.move_to(destination)
                    #return

                await asyncio.sleep(2)
                message_queue = deque([])
                can_speak = True
                message = name
                usernick = ctx.message.author.display_name
                message = message
                try:
                    vc = ctx.message.guild.voice_client
                    if not vc.is_playing():
                        tts = gTTS(message)
                        f = TemporaryFile()
                        tts.write_to_fp(f)
                        f.seek(0)
                        vc.play(discord.FFmpegPCMAudio(f,pipe=True,executable=r"C:\SAMEER MAIN\Sameer\Code\Lemen Bot\LemenBot\ffmpeg\bin\ffmpeg.exe"))
        
                    else:
                        message_queue.append(message)
                        if vc.is_playing():
                            await asyncio.sleep(10)
                        tts = gTTS(message_queue.popleft())
                        f = TemporaryFile()
                        tts.write_to_fp(f)
                        f.seek(0)
                        vc.play(discord.FFmpegPCMAudio(f, pipe=True,executable=r"C:\SAMEER MAIN\Sameer\Code\Lemen Bot\LemenBot\ffmpeg\bin\ffmpeg.exe"))
                except(TypeError, AttributeError):
                    try:
                        tts = gTTS(message)
                        f = TemporaryFile()
                        tts.write_to_fp(f)
                        f.seek(0)
                        channel = destination
                        vc = await channel.connect()
                        vc.play(discord.FFmpegPCMAudio(f, pipe=True,executable=r"C:\SAMEER MAIN\Sameer\Code\Lemen Bot\LemenBot\ffmpeg\bin\ffmpeg.exe"))
                    except(AttributeError, TypeError):
                        await ctx.send("I'm not in a voice channel and neither are you!")
                    return
                f.close()
                await asyncio.sleep(10)
                await ctx.voice_client.disconnect(force=True)



    @_join.before_invoke
    @_play.before_invoke
    async def ensure_voice_state(self, ctx: commands.Context):
        if not ctx.author.voice or not ctx.author.voice.channel:
            raise commands.CommandError('You are not connected to any voice channel.')

        if ctx.voice_client:
            if ctx.voice_client.channel != ctx.author.voice.channel:
                destination = ctx.author.voice.channel
                if ctx.voice_state.voice:
                    await ctx.voice_state.voice.move_to(destination)
                    return
                ctx.voice_state.voice = await destination.connect()

def setup(client):
    client.add_cog(Music(client))